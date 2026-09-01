"""Coverage for table detection and editable spreadsheet reconstruction."""

from __future__ import annotations

import os
import sys
import tempfile
import unittest
from types import SimpleNamespace
from unittest.mock import patch

import fitz
from openpyxl import Workbook, load_workbook

SERVER_DIR = os.path.dirname(os.path.dirname(__file__))
if SERVER_DIR not in sys.path:
    sys.path.insert(0, SERVER_DIR)

from pdf_export_model import TextWord
from pdf_export_spreadsheet import (
    export_xlsx,
    native_tables,
    row_segments,
    safe_sheet_name,
    style_sheet,
    words_to_grid,
)


def word(text: str, x0: float, y0: float, x1: float, y1: float) -> TextWord:
    return TextWord(text, x0, y0, x1, y1)


def report(pdf_type: str = "native") -> dict:
    return {
        "pdfType": pdf_type,
        "tablesDetected": 0,
        "editableTextBoxes": 0,
        "editableCharacters": 0,
        "editableTextDetected": False,
        "ocrTextCandidates": 0,
        "lowConfidenceOcrAreas": 0,
        "ocrPasses": [],
        "textCoverageEstimate": 0,
        "nonEditableVisualFallback": False,
        "warnings": [],
        "notes": [],
    }


def save_pdf(path: str, rows: list[tuple[str, float, float]] | None = None) -> None:
    document = fitz.open()
    page = document.new_page(width=300, height=180)
    for text, x, y in rows or []:
        page.insert_text((x, y), text, fontsize=10)
    document.save(path)
    document.close()


class SpreadsheetModelTests(unittest.TestCase):
    def test_native_tables_normalizes_cells_and_drops_empty_rows(self) -> None:
        table = SimpleNamespace(
            extract=lambda: [[" Name ", " Value\x00 "], [" ", None]]
        )
        page = SimpleNamespace(find_tables=lambda: SimpleNamespace(tables=[table]))

        self.assertEqual(native_tables(page), [[["Name", "Value"]]])

    def test_native_tables_returns_empty_on_detector_failure(self) -> None:
        page = SimpleNamespace(
            find_tables=lambda: (_ for _ in ()).throw(RuntimeError())
        )
        self.assertEqual(native_tables(page), [])

    def test_row_segments_splits_large_gaps_and_joins_close_words(self) -> None:
        words = [
            word("A", 0, 10, 10, 20),
            word("one", 13, 10, 30, 20),
            word("B", 100, 10, 110, 20),
        ]
        self.assertEqual(row_segments(words), ["A one", "B"])
        self.assertEqual(row_segments([]), [])

    def test_words_to_grid_groups_visual_rows_and_columns(self) -> None:
        words = [
            word("B2", 100, 51, 115, 61),
            word("A1", 0, 10, 15, 20),
            word("B1", 100, 11, 115, 21),
            word("A2", 0, 50, 15, 60),
        ]
        self.assertEqual(words_to_grid(words), [["A1", "B1"], ["A2", "B2"]])
        self.assertEqual(words_to_grid([]), [])

    def test_safe_sheet_name_removes_illegal_chars_and_bounds_length(self) -> None:
        self.assertEqual(safe_sheet_name(" []:*?/\\ "), "Sheet")
        self.assertEqual(len(safe_sheet_name("x" * 40)), 31)
        self.assertEqual(safe_sheet_name("Page: 1/Table"), "Page 1Table")

    def test_style_sheet_formats_headers_and_bounds_column_width(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["Header", "x" * 100])
        worksheet.append(["Body", "value"])

        style_sheet(worksheet, max_row=2, max_col=2)

        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertTrue(worksheet["A1"].font.bold)
        self.assertEqual(worksheet["A1"].fill.fgColor.rgb, "00EAF2FF")
        self.assertEqual(worksheet.column_dimensions["A"].width, 10)
        self.assertEqual(worksheet.column_dimensions["B"].width, 38)

    def test_export_xlsx_reconstructs_positioned_text_and_updates_report(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            source = os.path.join(tempdir, "source.pdf")
            output = os.path.join(tempdir, "output.xlsx")
            save_pdf(
                source,
                [
                    ("A1", 10, 30),
                    ("B1", 180, 30),
                    ("A2", 10, 60),
                    ("B2", 180, 60),
                    ("A3", 10, 90),
                    ("B3", 180, 90),
                ],
            )
            metrics = report()

            export_xlsx(source, output, metrics, "auto", table_detection=False)

            workbook = load_workbook(output)
            worksheet = workbook["Page 1"]
            self.assertEqual(worksheet["A1"].value, "A1")
            self.assertEqual(worksheet["B3"].value, "B3")
            self.assertEqual(metrics["tablesDetected"], 1)
            self.assertEqual(metrics["editableTextBoxes"], 6)
            self.assertTrue(metrics["editableTextDetected"])
            self.assertIn("inferred editable table", metrics["notes"][0])
            workbook.close()

    @patch("pdf_export_spreadsheet.native_tables")
    def test_export_xlsx_prefers_detected_native_tables(self, find_tables) -> None:
        find_tables.return_value = [[["Name", "Value"], ["File", "Mint"]]]
        with tempfile.TemporaryDirectory() as tempdir:
            source = os.path.join(tempdir, "source.pdf")
            output = os.path.join(tempdir, "output.xlsx")
            save_pdf(source)
            metrics = report()

            export_xlsx(source, output, metrics, "auto", table_detection=True)

            workbook = load_workbook(output)
            worksheet = workbook["Page 1 Table 1"]
            self.assertEqual(worksheet["B2"].value, "Mint")
            self.assertEqual(metrics["tablesDetected"], 1)
            self.assertIn("rebuilt 1 native PDF table", metrics["notes"][0])
            workbook.close()

    def test_export_xlsx_marks_blank_pages_as_visual_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as tempdir:
            source = os.path.join(tempdir, "blank.pdf")
            output = os.path.join(tempdir, "blank.xlsx")
            save_pdf(source)
            metrics = report()

            export_xlsx(source, output, metrics, "auto", table_detection=False)

            workbook = load_workbook(output)
            self.assertIn("No extractable text", workbook.active["A1"].value)
            self.assertTrue(metrics["nonEditableVisualFallback"])
            self.assertEqual(metrics["textCoverageEstimate"], 0)
            self.assertIn("Page 1", metrics["warnings"][0])
            workbook.close()


if __name__ == "__main__":
    unittest.main()
