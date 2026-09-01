"""Editable table and spreadsheet reconstruction for PDF exports."""

from __future__ import annotations

import statistics
import tempfile
from typing import Any

import fitz

from pdf_export_model import TextWord, clean_text
from pdf_export_text import join_positioned_words, maybe_resolve_ocr, page_text_words


def native_tables(page: fitz.Page) -> list[list[list[str]]]:
    """Extract and normalize tables discovered by PyMuPDF."""
    tables: list[list[list[str]]] = []
    try:
        found = page.find_tables()
        for table in getattr(found, "tables", []) or []:
            data = table.extract()
            cleaned = [
                [clean_text(cell) for cell in row]
                for row in data
                if any(clean_text(cell) for cell in row)
            ]
            if cleaned:
                tables.append(cleaned)
    except Exception:
        return []
    return tables


def row_segments(words: list[TextWord]) -> list[str]:
    """Split one visual row into cells using unusually large horizontal gaps."""
    if not words:
        return []
    words = sorted(words, key=lambda word: word.x0)
    heights = [max(1.0, word.y1 - word.y0) for word in words]
    median_height = statistics.median(heights) if heights else 10.0
    segments: list[list[TextWord]] = [[words[0]]]
    last_right = words[0].x1
    for word in words[1:]:
        gap = word.x0 - last_right
        if gap > max(18.0, median_height * 1.7):
            segments.append([word])
        else:
            segments[-1].append(word)
        last_right = max(last_right, word.x1)
    return [text for segment in segments if (text := join_positioned_words(segment))]


def words_to_grid(words: list[TextWord]) -> list[list[str]]:
    """Group positioned words into editable rows and inferred cells."""
    if not words:
        return []
    ordered = sorted(words, key=lambda word: (word.y0, word.x0))
    heights = [max(1.0, word.y1 - word.y0) for word in ordered]
    tolerance = max(4.0, (statistics.median(heights) if heights else 10.0) * 0.7)
    rows: list[list[TextWord]] = []
    centers: list[float] = []
    for word in ordered:
        center_y = (word.y0 + word.y1) / 2.0
        row_index = None
        for index, center in enumerate(centers):
            if abs(center_y - center) <= tolerance:
                row_index = index
                break
        if row_index is None:
            rows.append([word])
            centers.append(center_y)
        else:
            rows[row_index].append(word)
            centers[row_index] = statistics.mean(
                [(row_word.y0 + row_word.y1) / 2.0 for row_word in rows[row_index]]
            )
    grid = [row_segments(row) for row in rows]
    return [row for row in grid if any(clean_text(cell) for cell in row)]


def style_sheet(worksheet: Any, max_row: int, max_col: int) -> None:
    """Apply a readable, bounded spreadsheet style to an exported worksheet."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="C9D3DF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True, color="1F2937")
                cell.fill = header_fill
    for column in range(1, max_col + 1):
        letter = get_column_letter(column)
        longest = 0
        for row in range(1, max_row + 1):
            value = worksheet.cell(row=row, column=column).value
            longest = max(longest, len(str(value or "")))
        worksheet.column_dimensions[letter].width = max(10, min(38, longest + 2))
    worksheet.freeze_panes = "A2"


def safe_sheet_name(name: str) -> str:
    """Return a legal Excel worksheet name of at most 31 characters."""
    cleaned = (
        "".join(char for char in name if char not in r"[]:*?/\\").strip() or "Sheet"
    )
    return cleaned[:31]


def export_xlsx(
    src: str,
    dst: str,
    report: dict[str, Any],
    lang: str,
    table_detection: bool,
) -> None:
    """Export native tables or positioned text into an editable workbook."""
    from openpyxl import Workbook

    doc = fitz.open(src)
    workbook = Workbook()
    workbook.remove(workbook.active)
    ocr_lang = maybe_resolve_ocr(src, lang, True, report)
    table_count = 0
    inferred_table_count = 0
    cells_written = 0
    chars_written = 0

    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            for page in doc:
                page_tables = native_tables(page) if table_detection else []
                if page_tables:
                    for index, table in enumerate(page_tables, start=1):
                        table_count += 1
                        worksheet = workbook.create_sheet(
                            safe_sheet_name(f"Page {page.number + 1} Table {index}")
                        )
                        for row_index, row in enumerate(table, start=1):
                            for column_index, value in enumerate(row, start=1):
                                worksheet.cell(row_index, column_index, value)
                                if value:
                                    cells_written += 1
                                    chars_written += len(value)
                        style_sheet(
                            worksheet,
                            len(table),
                            max(len(row) for row in table),
                        )
                    continue

                words = page_text_words(page, tmpdir, ocr_lang, report, allow_ocr=True)
                grid = words_to_grid(words)
                worksheet = workbook.create_sheet(
                    safe_sheet_name(f"Page {page.number + 1}")
                )
                if not grid:
                    worksheet["A1"] = (
                        "No extractable text or table structure was detected on this page."
                    )
                    report["warnings"].append(
                        f"Page {page.number + 1} exported without editable table/text cells."
                    )
                    report["nonEditableVisualFallback"] = True
                    continue
                for row_index, row in enumerate(grid, start=1):
                    for column_index, value in enumerate(row, start=1):
                        worksheet.cell(row_index, column_index, value)
                        if value:
                            cells_written += 1
                            chars_written += len(value)
                max_columns = max(len(row) for row in grid)
                if len(grid) >= 3 and max_columns >= 2:
                    inferred_table_count += 1
                style_sheet(worksheet, len(grid), max_columns)
        finally:
            doc.close()

    if not workbook.worksheets:
        workbook.create_sheet("Result")
        workbook.active["A1"] = "No pages were converted."

    report["tablesDetected"] = max(
        int(report.get("tablesDetected") or 0), table_count + inferred_table_count
    )
    report["editableTextBoxes"] = cells_written
    report["editableCharacters"] = chars_written
    report["editableTextDetected"] = cells_written > 0
    report["textCoverageEstimate"] = 100 if cells_written else 0
    if table_count:
        report["notes"].append(
            f"Excel export rebuilt {table_count} native PDF table(s) as editable worksheets."
        )
    elif inferred_table_count:
        report["notes"].append(
            "Excel export inferred editable table-like worksheets from positioned PDF/OCR text."
        )
    elif cells_written:
        report["notes"].append(
            "Excel export grouped PDF text into editable rows and columns."
        )
    workbook.save(dst)
